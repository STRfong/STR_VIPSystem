from django.contrib.auth.decorators import login_required   
from django.utils.decorators import method_decorator
from django.views.generic import ListView, UpdateView
from VIPSystem_APP.models import ProjectParticipation, VIP, Project, EventTime, EventTicket, Tag
from django.shortcuts import get_object_or_404, redirect
from django.db import transaction
from django.contrib import messages
from django.db.models import Prefetch
from django.contrib.auth.models import User
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from django.shortcuts import get_object_or_404
from datetime import datetime
from django.utils.encoding import escape_uri_path
from django.http import HttpResponse
from django.http import JsonResponse
from django.views.decorators.http import require_POST





@method_decorator(login_required, name='dispatch')
class ProjectParticipantsView(ListView):
    model = ProjectParticipation
    template_name = 'VIPSystem/project_participants.html'
    context_object_name = 'participants_list'
    pk_url_kwarg = 'project_id'

    def get_queryset(self):
        project_id = self.kwargs.get('project_id')
        return ProjectParticipation.objects.filter(project_id= project_id)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project_id = self.kwargs.get('project_id')
        context['project'] = get_object_or_404(Project, pk=project_id)
        return context
    
@method_decorator(login_required, name='dispatch')
class ProjectParticipationBySectionView(ListView):
    model = ProjectParticipation
    template_name = 'VIPSystem/participation_by_section.html'
    context_object_name = 'participation_list'
    paginate_by = 20  # 每页显示10条记录

    def get_queryset(self):
        project = get_object_or_404(Project, id=self.kwargs['project_id'])
        queryset = ProjectParticipation.objects.filter(
            project=project, 
            wish_attend_section=self.kwargs['section']
        )

        name_filter = self.request.GET.get('nameFilter')
        if name_filter:
            queryset = queryset.filter(vip__name__icontains=name_filter)

        invited_by_filter = self.request.GET.getlist('filter_invited_by')
        if invited_by_filter:
            queryset = queryset.filter(invited_by_id__in=invited_by_filter)

        status_filter = self.request.GET.getlist('filter_status')
        if status_filter:
            queryset = queryset.filter(status__in=status_filter)

        filter_wish_attend = self.request.GET.getlist('filter_wish_attend')
        if filter_wish_attend:
            if 'all' in filter_wish_attend:
                event_times = EventTime.objects.filter(project_id=self.kwargs['project_id'], section=self.kwargs['section'])
            else:
                event_times = EventTime.objects.filter(id__in=filter_wish_attend)
            queryset = [ participation for participation in queryset if self.check_intersection(event_times, participation)]
        
        
        return queryset
        
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project = Project.objects.get(id=self.kwargs['project_id'])
        context['project'] = project
        context['section'] = self.kwargs['section']
        context['event_times'] = EventTime.objects.filter(project_id=self.kwargs['project_id'], section=self.kwargs['section'])
        context['staffs'] = User.objects.all()
        context['username'] = self.request.user.username
        # context['dead_line_date'] = EventTime.objects.filter(project_id=self.kwargs['project_id'], section=self.kwargs['section']).first().dead_line_date
        return context
    
    def check_intersection(self, event_times, participation):
        participation_set = set(participation.get_wish_attend_list())
        event_times_set = set(event_times)
        return len(event_times_set.intersection(participation_set))>0
    
@method_decorator(login_required, name='dispatch')
class ProjectParticipantsByEventTimeView(ListView):
    model = ProjectParticipation
    template_name = 'VIPSystem/project_participants_event_time.html'
    context_object_name = 'participants_list'
    pk_url_kwarg = 'project_id'

    def get_queryset(self):
        project_id = self.kwargs.get('project_id')
        event_time = EventTime.objects.get(id=self.kwargs.get('event_time_id'))
        section = self.kwargs.get('section')
        queryset = ProjectParticipation.objects.filter(project_id=project_id, wish_attend_section=section).order_by('invited_by__username')
        
        # 篩選名字
        name_filter = self.request.GET.get('nameFilter')
        if name_filter:
            queryset = queryset.filter(vip__name__icontains=name_filter)

        invited_by_filter = self.request.GET.getlist('filter_invited_by')
        if invited_by_filter:
            queryset = queryset.filter(invited_by_id__in=invited_by_filter)

        status_filter = self.request.GET.getlist('filter_status')
        if status_filter:
            queryset = queryset.filter(status__in=status_filter)

        organization_filter = self.request.GET.get('organizationFilter')
        if organization_filter:
            queryset = queryset.filter(vip__organization__icontains=organization_filter)

        return_queryset = [ participation for participation in queryset if self.check_intersection(event_time, participation)]
        return return_queryset
    
    def check_intersection(self, event_time, participation):
        participation_set = set(participation.get_wish_attend_list())
        if participation.status == 'added' or participation.status == 'sended' or participation.status == 'declined':
            return event_time in participation_set
        elif participation.status == 'confirmed':
            return event_time == participation.event_time
        else:
            return False

    def get_context_data(self, **kwargs):
        try:
            context = super().get_context_data(**kwargs)
            project_id = self.kwargs.get('project_id')
            event_time_id = self.kwargs.get('event_time_id')
            context['project'] = get_object_or_404(Project, pk=project_id)
            context['event_time'] = get_object_or_404(EventTime, pk=event_time_id)
            context['event_times'] = EventTime.objects.filter(project_id=self.kwargs['project_id'], section=self.kwargs['section'])
            context['event_times_json'] = EventTime.get_event_times_json(
                self.kwargs['project_id'], 
                self.kwargs['section']
            )
            context['staffs'] = User.objects.all()
            context['username'] = self.request.user.username
            context['invited_amount'] = ProjectParticipation.objects.filter(project_id=project_id, event_time_id=event_time_id, invited_by=self.request.user).count()
            try:
                event_ticket = EventTicket.objects.get(
                    event_time_id=event_time_id, 
                    staff__user=self.request.user
                )
                context['ticket_count'] = event_ticket.ticket_count
                context['invited_ticket_count'] = event_ticket.total_invited_tickets()
                context['remaining_ticket_count'] = event_ticket.total_remaining_tickets()
            except EventTicket.DoesNotExist:
                context['ticket_count'] = 0
                context['invited_ticket_count'] = 0
                context['remaining_ticket_count'] = 0
            
            # 取得所有同仁的邀請狀況
            all_ticket_status = self.get_all_staff_invite_status(event_time_id)
            context['staff_invite_status'] = all_ticket_status['staff_invite_status']
            context['total_invited_tickets'] = all_ticket_status['total_invited_tickets']
            return context
        except Exception as e:
            print(f"Error: {e}")
            return context
        
    def get_all_staff_invite_status(self, event_time_id):
        staff_invite_status = {}
        staffs = User.objects.all()
        for staff in staffs:
            try:
                event_ticket = EventTicket.objects.get(
                    event_time_id=event_time_id, 
                    staff__user=staff
                )
                staff_invite_status[staff.profile.nickname] = {"可使用票數": event_ticket.ticket_count, "已使用票數": event_ticket.total_invited_tickets(), "剩餘票數": event_ticket.total_remaining_tickets()}
            except EventTicket.DoesNotExist:
                # staff_invite_status[staff.profile.nickname] = {"可使用票數": 0, "已使用票數": 0, "剩餘票數": 0}
                continue
        
        # 計算所有同仁的票數總和
        total_ticket_count = sum(status["可使用票數"] for status in staff_invite_status.values())
        total_invited_tickets = sum(status["已使用票數"] for status in staff_invite_status.values())  
        total_remaining_tickets = sum(status["剩餘票數"] for status in staff_invite_status.values())

        # 將總和加入到 staff_invite_status 中
        total_invited_tickets = {
            "可使用票數": total_ticket_count,
            "已使用票數": total_invited_tickets,
            "剩餘票數": total_remaining_tickets
        }
        return {'staff_invite_status': staff_invite_status, 'total_invited_tickets': total_invited_tickets}
    
@method_decorator(login_required, name='dispatch') # 邀請貴賓參與專案
class InviteListView(ListView):
    model = VIP
    template_name = 'VIPSystem/invite_list.html'
    context_object_name = 'vip_list'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project_id = self.kwargs.get('project_id')
        project = get_object_or_404(Project, pk=project_id)
        context['project'] = project
        context['current_participants'] = project.participants.all()
        return context
    
@method_decorator(login_required, name='dispatch') # 邀請貴賓參與專案
class InviteListBySectionView(ListView):
    model = VIP
    template_name = 'VIPSystem/invite_list_by_section.html'
    context_object_name = 'vip_list'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project_id = self.kwargs.get('project_id')
        section = self.kwargs.get('section')
        project = get_object_or_404(Project, pk=project_id)
        context['project'] = project
        context['current_participants'] = project.participants.all() # 可能要修改一下
        context['section'] = section
        context['event_times'] = EventTime.objects.filter(project_id=project_id, section=section)
        return context
    
@method_decorator(login_required, name='dispatch') # 邀請貴賓參與專案
class InviteListViewEventTime(ListView):
    model = VIP
    template_name = 'VIPSystem/invite_list_event_time.html'
    context_object_name = 'vip_list'

    def get_queryset(self):
        queryset = super().get_queryset()
        name_filter = self.request.GET.get('nameFilter')
        if name_filter:
            queryset = queryset.filter(name__icontains=name_filter)
        organization_filter = self.request.GET.get('organizationFilter')
        if organization_filter:
            queryset = queryset.filter(organization__icontains=organization_filter)
        position_filter = self.request.GET.get('positionFilter')
        if position_filter:
            queryset = queryset.filter(position__icontains=position_filter)
        tag_filter = self.request.GET.getlist('tags')
        if tag_filter:
            queryset = queryset.filter(tags__in=tag_filter)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project_id = self.kwargs.get('project_id')
        event_time_id = self.kwargs.get('event_time_id')
        project = get_object_or_404(Project, pk=project_id)
        context['project'] = project
        context['current_participants'] = project.participants.all()
        context['all_tags'] = Tag.objects.all()
        context['event_time'] = get_object_or_404(EventTime, pk=event_time_id)
        return context

@method_decorator(login_required, name='dispatch') # 邀請貴賓參與專案
class UpdateParticipantsBySectionView(UpdateView):
    def post(self, request, *args, **kwargs):        
        project_id = kwargs.get('project_id')
        project = Project.objects.get(pk=project_id)
        section = kwargs.get('section')
        vip_list = request.POST.getlist('selected_vips')
        event_time_list = request.POST.getlist('selected_event_time_by_section')
        for vip_id in vip_list:
            vip = VIP.objects.get(id=vip_id)
            ProjectParticipation.objects.update_or_create(
                project=project,
                vip=vip,
                invited_by=request.user,
                status = 'added',
                wish_attend = self.event_time_selected(event_time_list),
                wish_attend_section=section
            )  
        messages.success(request, f'已新增 {len(vip_list)} 位貴賓至 {section} 場。')
        return redirect('VIPSystem_APP:participation_by_section', project_id=project_id, section=section)
    
    def event_time_selected(self, event_time_list):
        if 'all' in event_time_list:
            return 'all'
        else:
            return ','.join(event_time_list)
        
@method_decorator(login_required, name='dispatch') # 邀請貴賓參與專案
class UpdateParticipantsInfoBySectionView(UpdateView):
    def post(self, request, *args, **kwargs):
        project_id = kwargs.get('project_id')
        section = kwargs.get('section')
        vip_id = request.POST.get('vip_id')
        vip = VIP.objects.get(id=vip_id)
        project_participation = ProjectParticipation.objects.get(project=project_id, vip=vip)
        project_participation.wish_attend = self.event_time_selected(request.POST.getlist('selected_event_time_by_section'))
        project_participation.wish_ticket_count = request.POST.get('wish_ticket_count')
        project_participation.save()
        return redirect('VIPSystem_APP:participation_by_section', project_id=project_id, section=section)
    
    def event_time_selected(self, event_time_list):
        if 'all' in event_time_list:
            return 'all'
        else:
            return ','.join(event_time_list)
        
@method_decorator(login_required, name='dispatch')
class UpdateParticipantsByEventTimeView(UpdateView):
    def post(self, request, *args, **kwargs):
        project_id = kwargs.get('project_id')
        section = kwargs.get('section')
        event_time_id = kwargs.get('event_time_id')
        
        project = get_object_or_404(Project, id=project_id)
        event_time = get_object_or_404(EventTime, id=event_time_id)
        
        selected_vips = request.POST.getlist('selected_vips')
        
        # 添加選中的VIP到參與者列表
        for vip_id in selected_vips:
            vip = get_object_or_404(VIP, id=vip_id)
            ProjectParticipation.objects.update_or_create(
                project=project,
                vip=vip,
                invited_by=request.user,
                status='added',
                event_time=event_time,
                wish_attend_section=section,
                wish_attend=str(event_time.id)
            )
        
        messages.success(request, f'已新增 {len(selected_vips)} 位貴賓至 {event_time} 場次。')
        return redirect('VIPSystem_APP:participation_by_event_time', 
                       project_id=project_id, 
                       section=section, 
                       event_time_id=event_time_id)

    def get(self, request, *args, **kwargs):
        # 如果是GET請求，重定向回邀請列表頁面
        return redirect('VIPSystem_APP:invite_list_event_time', 
                       project_id=kwargs.get('project_id'), 
                       event_time_id=kwargs.get('event_time_id'))
        
@method_decorator(login_required, name='dispatch')
class UpdateParticipantsByEventTimeDirectlyView(UpdateView):
    def post(self, request, *args, **kwargs):
        project_id = kwargs.get('project_id')
        section = kwargs.get('section')
        event_time_id = kwargs.get('event_time_id')
        # 獲取表單數據
        vip_name = request.POST.get('name')
        vip_nickname = request.POST.get('nickname')
        vip_organization = request.POST.get('organization')
        vip_position = request.POST.get('position')
        vip_phone = request.POST.get('phone')
        vip_email = request.POST.get('email')
        str_connect_name = request.POST.get('str_connect')
        try:
            str_connect = User.objects.get(profile__nickname=str_connect_name)
        except User.DoesNotExist:
            messages.error(request, f'加入失敗：找不到叫 {str_connect_name} 的同仁ㄛ，試著在下拉選單中找找看！')
            return redirect('VIPSystem_APP:participation_by_event_time', 
                       project_id=project_id, 
                       section=section, 
                       event_time_id=event_time_id)

        vip_notes = request.POST.get('vip_notes')
        wish_ticket_count = request.POST.get('wish_ticket_count')
        poc = request.POST.get('poc')
        poc_position = request.POST.get('poc_position')
        poc_phone_number = request.POST.get('poc_phone_number')
        poc_email = request.POST.get('poc_email')
        pp_notes = request.POST.get('pp_notes')
        
        try:
            with transaction.atomic():
                # 嘗試查找或創建 VIP
                vip = VIP.objects.filter(email=vip_email).first()
                
                if vip:
                    # 更新現有 VIP
                    messages.info(request, f'找到現有貴賓：{vip.name}')
                    vip.name = vip_name
                    vip.nickname = vip_nickname
                    vip.email = vip_email
                    vip.phone_number = vip_phone
                    vip.organization = vip_organization
                    vip.position = vip_position
                    vip.str_connect = str_connect.profile.nickname
                    vip.notes = vip_notes
                    vip.poc = poc
                    vip.poc_position = poc_position
                    vip.poc_phone_number = poc_phone_number
                    vip.poc_email = poc_email
                    vip.save()
                else:
                    # 創建新的 VIP
                    vip = VIP.objects.create(
                        name=vip_name,
                        nickname=vip_nickname,
                        email=vip_email,
                        phone_number=vip_phone,
                        organization=vip_organization,
                        position=vip_position,
                        str_connect=str_connect.profile.nickname,
                        notes=vip_notes,
                        poc=poc,
                        poc_position=poc_position,
                        poc_phone_number=poc_phone_number,
                        poc_email=poc_email
                    )
                    messages.success(request, f'已創建新貴賓：{vip.name}')
                
                # 獲取項目和場次
                project = get_object_or_404(Project, id=project_id)
                event_time = get_object_or_404(EventTime, id=event_time_id)
                
                # 創建或更新參與記錄
                participation, created = ProjectParticipation.objects.update_or_create(
                    project=project,
                    vip=vip,
                    event_time=event_time,
                    invited_by=str_connect,
                    notes=pp_notes,
                    defaults={
                        'status': 'added',
                        'wish_attend': str(event_time.id),
                        'wish_attend_section': section,
                        'wish_ticket_count': wish_ticket_count,
                        'join_people_count': 0
                    }
                )
                
                if created:
                    messages.success(request, '已成功將貴賓加入本場次')
                else:
                    messages.info(request, '已更新貴賓的參與資訊')
                
        except Exception as e:
            messages.error(request, f'處理過程中發生錯誤：{str(e)}')
            
        return redirect('VIPSystem_APP:participation_by_event_time', 
                       project_id=project_id, 
                       section=section, 
                       event_time_id=event_time_id)
    def get(self, request, *args, **kwargs):
        return redirect('VIPSystem_APP:participation_by_event_time',
                       project_id=kwargs.get('project_id'),
                       section=kwargs.get('section'),
                       event_time_id=kwargs.get('event_time_id'))
    
@method_decorator(login_required, name='dispatch')
class ExportParticipantsByEventTimeView(ListView):
    model = ProjectParticipation
    
    def get_queryset(self):
        # 獲取URL參數
        project_id = self.kwargs.get('project_id')
        event_time_id = self.kwargs.get('event_time_id')
        
        # 獲取並驗證專案
        self.project = get_object_or_404(Project, pk=project_id)
        
        # 獲取並驗證場次屬於該專案
        self.event_time = get_object_or_404(
            EventTime, 
            id=event_time_id,
            project=self.project
        )
        
        # 返回過濾後的查詢集
        return ProjectParticipation.objects.filter(
            project=self.project,
            event_time=self.event_time
        ).select_related(
            'vip',
            'invited_by'
        ).order_by('pp_id')

    def get(self, request, *args, **kwargs):
        # 取得資料
        queryset = self.get_queryset()
        
        # 創建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = f"{self.project.name}-{self.event_time.date.strftime('%Y%m%d')}-{self.event_time.session}"

        # 設定標題樣式
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        # 定義欄位
        headers = [
            '取票序號', '姓名', '稱呼', '單位', '職稱', 
            '電話', 'Email', '聯絡窗口', '窗口職稱',
            '窗口電話', '窗口Email', '公司地址',
            '邀請人', '狀態', '參與人數', '備註', 
            '出席確認信', '出席提醒信'
        ]

        # 寫入標題
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # 寫入資料
        status_mapping = {
            'added': '尚未寄出邀請信',
            'confirmed': '確認參加',
            'sended': '已發送邀請信件等待回覆',
            'declined': '拒絕參加'
        }

        for row, pp in enumerate(queryset, 2):
            data = [
                pp.pp_id,
                pp.vip.name,
                pp.vip.nickname,
                pp.vip.organization,
                pp.vip.position,
                pp.vip.phone_number,
                pp.vip.email,
                pp.vip.poc,
                pp.vip.poc_position,
                pp.vip.poc_phone_number,
                pp.vip.poc_email,
                pp.vip.address,
                pp.invited_by.username if pp.invited_by else '',
                status_mapping.get(pp.status, pp.status),
                pp.join_people_count,
                pp.notes,
                '⭕️' if pp.send_check_email else '❌',
                '⭕️' if pp.send_remind_email else '❌'
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)

        # 調整欄寬
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 準備檔案回應
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{self.project.name}_{self.event_time.date.strftime('%Y%m%d')}_{self.event_time.session}_貴賓名單.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{escape_uri_path(filename)}"'

        # 儲存到回應
        wb.save(response)
        return response

@method_decorator(login_required, name='dispatch') # 邀請貴賓參與專案
class UpdateParticipantsInfoByEventTimeView(UpdateView):
    def post(self, request, *args, **kwargs):
        project_id = kwargs.get('project_id')
        event_time_id = kwargs.get('event_time_id')
        section = kwargs.get('section')
        vip_id = request.POST.get('vip_id')
        vip = VIP.objects.get(id=vip_id)
        project_participation = ProjectParticipation.objects.get(project=project_id, vip=vip)
        # project_participation.wish_attend = self.event_time_selected(request.POST.getlist('selected_event_time_by_section')) # 暫時先屏蔽，之後再打開
        project_participation.wish_ticket_count = request.POST.get('wish_ticket_count')
        project_participation.save()
        messages.success(request, f'已更新 {vip.name} 的希望參加場次和提供票數')
        return redirect('VIPSystem_APP:participation_by_event_time', project_id=project_id, section=section, event_time_id=event_time_id)
    
    def event_time_selected(self, event_time_list):
        if 'all' in event_time_list:
            return 'all'
        else:
            return ','.join(event_time_list)
        

    
@method_decorator(login_required, name='dispatch') # 發送邀請信
class SendEmailListBySectionView(ListView):
    model = ProjectParticipation
    template_name = 'VIPSystem/send_emails_by_section.html'
    context_object_name = 'participants_list'

    def get_queryset(self):
        project_id = self.kwargs.get('project_id')
        return ProjectParticipation.objects.filter(project_id=project_id).exclude(status='confirmed')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project_id = self.kwargs.get('project_id')
        section = self.kwargs.get('section')
        project = get_object_or_404(Project, pk=project_id)
        context['project'] = project
        context['section'] = section
        context['event_times'] = EventTime.objects.filter(project_id=project_id, section=section)
        return context
    
@method_decorator(login_required, name='dispatch') # 發送邀請信
class SendEmailListViewEventTime(ListView):
    model = ProjectParticipation
    template_name = 'VIPSystem/send_emails_event_time.html'
    context_object_name = 'vip_list'

    def get_queryset(self):
        project_id = self.kwargs.get('project_id')
        event_time_id = self.kwargs.get('event_time_id')
        event_time = get_object_or_404(EventTime, id=event_time_id)
        
        queryset = ProjectParticipation.objects.filter(project_id=project_id)
        filtered_queryset = [
            participation for participation in queryset 
            if event_time in participation.get_wish_attend_list()
        ]
    
        return [participation for participation in filtered_queryset if participation.status != 'confirmed']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['project'] = get_object_or_404(Project, pk=self.kwargs.get('project_id'))
        context['event_time'] = get_object_or_404(EventTime, pk=self.kwargs.get('event_time_id'))
        return context
    
@login_required
def update_participants(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    
    if request.method == 'POST':
        selected_vips = set(map(int, request.POST.getlist('selected_vips')))
        # 添加選中的VIP到參與者列表
        for vip_id in selected_vips:
            vip = VIP.objects.get(id=vip_id)
            ProjectParticipation.objects.update_or_create(
                project=project,
                vip=vip,
                invited_by=request.user,
                status = 'added'
            )
        
        return redirect('VIPSystem_APP:project_participants', project_id=project.pk)
    
    # 如果不是POST請求，重定向回邀請列表頁面
    return redirect('VIPSystem_APP:invite_list', project_id=project.pk)


@login_required
def remove_participant(request, project_id, participant_id):
    if request.method == 'POST':
        project = get_object_or_404(Project, pk=project_id)
        participant = get_object_or_404(VIP, pk=participant_id)
        project_participation = get_object_or_404(ProjectParticipation, project=project, vip=participant)
        project_participation.delete()
        messages.success(request, f'已成功將 {participant.name} 從專案中移除。')
    return redirect('VIPSystem_APP:project_participants', project_id=project_id)

@login_required
def remove_participant_by_section(request, project_id, section):
    if request.method == 'POST':
        project = get_object_or_404(Project, pk=project_id)
        participant = get_object_or_404(VIP, pk=request.POST.get('participant_id'))
        project_participation = get_object_or_404(ProjectParticipation, project=project, vip=participant)
        project_participation.delete()
        messages.success(request, f'已成功將 {participant.name} 從專案中移除。')
    # 使用 HTTP_REFERER 返回到觸發此 function 的 URL
    referer = request.META.get('HTTP_REFERER')
    if referer:
        return redirect(referer)
    else:
        return redirect('VIPSystem_APP:participation_by_section', project_id=project_id, section=section)

@login_required
def remove_participant_event_time(request, project_id, section, event_time_id):
    if request.method == 'POST':
        project = get_object_or_404(Project, pk=project_id)
        event_time = get_object_or_404(EventTime, pk=event_time_id)
        participant = get_object_or_404(VIP, pk=request.POST.get('participant_id'))
        project_participation = get_object_or_404(ProjectParticipation, project=project, vip=participant, event_time=event_time)
        project_participation.delete()
        messages.success(request, f'已成功將 {participant.name} 從專案中移除。')
    return redirect('VIPSystem_APP:participation_by_event_time', project_id=project_id, section=section, event_time_id=event_time_id)

@require_POST
@login_required
def update_get_ticket_check(request, project_id, section, event_time_id):
    try:
        is_checked = request.POST.get('is_checked') == 'true'
        participant_id = request.POST.get('participant_id')
        
        participant = ProjectParticipation.objects.get(
            project_id=project_id,
            vip_id=participant_id,
            event_time_id=event_time_id
        )
        
        participant.get_ticket_check = is_checked
        participant.save()  # 這會觸發 signal
        
        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@require_POST
@login_required
def update_seat_number(request, project_id, section, event_time_id):
    try:
        # 獲取參數
        participant_id = request.POST.get('participant_id')
        seat_number = request.POST.get('seat_number', '').strip()

        # 獲取並驗證參與者記錄是否存在
        participant = ProjectParticipation.objects.get(
            vip_id=participant_id,
            event_time_id=event_time_id
        )

        # 更新座位號
        participant.seat_number = seat_number
        participant.save()  # 這會觸發 signal，自動處理 WebSocket 廣播

        return JsonResponse({'status': 'success'})

    except ProjectParticipation.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': '找不到該參與者記錄'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)